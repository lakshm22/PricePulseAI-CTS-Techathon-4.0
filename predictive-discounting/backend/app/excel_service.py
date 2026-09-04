import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

DEFAULT_HEADERS = [
    "SKU", "Barcode", "Product", "Category", "MRP",
    "Current Price", "Stock", "Expiry Date", "Discount %", "Updated At"
]

def get_excel_path() -> str:
    return os.getenv("EXCEL_FILE_PATH", "excel/store_inventory.xlsx")

def _ensure_workbook(path: str):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        ws.append(DEFAULT_HEADERS)
        wb.save(path)

def update_product_price(product, inventory, discount_percentage: float) -> dict:
    path = get_excel_path()
    _ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb["Inventory"] if "Inventory" in wb.sheetnames else wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    for header in DEFAULT_HEADERS:
        if header not in headers:
            col = ws.max_column + 1
            ws.cell(1, col, header)
            headers[header] = col

    row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, headers["SKU"]).value == product.sku:
            row = r
            break
    if row is None:
        row = ws.max_row + 1

    price = round(min(float(product.mrp), float(product.mrp) * (1 - discount_percentage / 100)), 2)
    values = {
        "SKU": product.sku, "Barcode": product.barcode, "Product": product.name,
        "Category": product.category, "MRP": float(product.mrp),
        "Current Price": price, "Stock": inventory.stock_quantity,
        "Expiry Date": product.expiry_date, "Discount %": discount_percentage,
        "Updated At": datetime.now()
    }
    for key, value in values.items():
        ws.cell(row, headers[key], value)
    wb.save(path)
    return {"file_path": path, "updated_price": price}
